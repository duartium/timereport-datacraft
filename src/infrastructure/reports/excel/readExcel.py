import openpyxl 
from io import BytesIO
from application.services.date_service import obtener_numero_mes
from datetime import datetime
from infrastructure.external_services.ipmAPI_service import post_importar_excel
import json
import locale
from application.services.date_service import get_days_of_month
from infrastructure.reports.excel.utils import agregar_acentos
def leer_excel(contenido, token):
    locale.setlocale(locale.LC_TIME, 'es_ES.utf8')
    wb = openpyxl.load_workbook(BytesIO(contenido), read_only=True)
    hoja = wb.active

    mes = hoja.cell(row=2, column=7).value
    if mes is None:
         mes = hoja.cell(row=2, column=6).value
    mes = obtener_numero_mes(mes)
    year = hoja.cell(row=2, column=10).value
    if year is None:
         year = hoja.cell(row=2, column=9).value
    fila = 8
    info_excel = {} 
    actividades =  []
    dias_mes = get_days_of_month(mes)
    while hoja.cell(row=fila, column=5).value is not None: 
        if str(hoja.cell(row=fila, column=1).value).isdigit():
                    fila_valores = [hoja.cell(row=fila, column=col).value for col in range(1, hoja.max_column + 1)]    
                    codigo_proyecto = fila_valores[3]    
                    if "_x000D_" in codigo_proyecto:
                        codigo_proyecto = fila_valores[3].replace("_x000D_", "").replace("\n", "")
                    if fila_valores[4] is None:
                        break
                    descripcion = str(fila_valores[4])
                    tipo_actividad = fila_valores[1].title().capitalize()
                    tipo_actividad_acc = agregar_acentos(tipo_actividad)
                    if tipo_actividad_acc is not None:
                          tipo_actividad = tipo_actividad_acc
                    actividad = {"tipoActividad":tipo_actividad, "codigoProyecto": codigo_proyecto, "descripcion": descripcion}
                    dias = fila_valores[6:-1] #columnas de dias
                    if len(dias_mes) < len(dias):
                         dias = fila_valores[6:-3]
                    arr_fechas_horas = []
                    i = 1                    
                    for dia in dias:
                        if dia is not None:
                            fecha = datetime(year, mes, i).isoformat() + "Z"
                            fecha_hora = {"fecha": fecha, "horas": dia}
                            arr_fechas_horas.append(fecha_hora)
                        i += 1
                    actividad["fechasHoras"] = arr_fechas_horas
                    actividades.append(actividad)
        fila += 1
    consultor =  hoja.cell(row=4, column=3).value
    if consultor is None:
         raise Exception("No existe nombre del consultor dentro del archivo Excel.")
    info_excel["usuario"] = consultor
    info_excel["actividades"] = actividades
    info_excel = eliminar_saltos_de_linea(info_excel)
    json_data = json.dumps(info_excel,ensure_ascii= False, indent=2)
    json_data = json.loads(json_data)
    res = post_importar_excel(json_data, token)
    return res


def eliminar_saltos_de_linea(diccionario):
    for clave, valor in diccionario.items():
        if isinstance(valor, str):
            diccionario[clave] = valor.replace('\n', '').replace('\r\n', '')
        elif isinstance(valor, dict):
            eliminar_saltos_de_linea(valor)
    return diccionario

